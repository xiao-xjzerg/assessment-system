"""最终考核成绩路由（M10）"""
import io
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import (
    ROLE_ADMIN, ROLE_LEADER, ROLE_PM, ROLE_EMPLOYEE,
    ASSESS_TYPE_MANAGER, ASSESS_TYPE_PUBLIC, ASSESS_TYPE_BUSINESS, ASSESS_TYPE_RD,
    RATING_LEVELS,
)
from app.database import get_db
from app.dependencies import get_current_user, require_roles
from app.models.employee import Employee
from app.models.cycle import Cycle
from app.schemas.result import FinalResultOut, RatingUpdate, LeaderCommentUpdate
from app.schemas.common import ResponseModel
from app.services.result_service import (
    calculate_final_results,
    get_final_results,
    update_rating,
    update_leader_comment,
)

router = APIRouter(prefix="/api/results", tags=["成绩汇总"])


async def _get_active_cycle(db: AsyncSession) -> Cycle:
    result = await db.execute(select(Cycle).where(Cycle.is_active == True))
    cycle = result.scalar_one_or_none()
    if cycle is None:
        raise HTTPException(status_code=400, detail="没有活跃的考核周期")
    return cycle


# ---- 触发最终成绩计算 ----
@router.post("/calculate", response_model=ResponseModel)
async def trigger_calculation(
    db: AsyncSession = Depends(get_db),
    current_user: Employee = Depends(require_roles([ROLE_ADMIN])),
):
    """管理员触发最终成绩全量计算（刷新计算）"""
    cycle = await _get_active_cycle(db)
    results = await calculate_final_results(db, cycle.id)
    data = [FinalResultOut.model_validate(r).model_dump() for r in results]
    return ResponseModel(message="成绩计算完成", data=data)


# ---- 查询最终成绩 ----
@router.get("", response_model=ResponseModel)
async def list_results(
    department: Optional[str] = Query(None),
    assess_type: Optional[str] = Query(None),
    employee_name: Optional[str] = Query(None),
    group_name: Optional[str] = Query(None),
    position: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    """
    查询最终考核成绩。
    管理员/领导可查全部，其他角色只查自己。
    """
    cycle = await _get_active_cycle(db)

    if current_user.role in (ROLE_ADMIN, ROLE_LEADER):
        items = await get_final_results(
            db, cycle.id,
            department=department,
            assess_type=assess_type,
            employee_name=employee_name,
            group_name=group_name,
            position=position,
        )
    else:
        items = await get_final_results(
            db, cycle.id,
            employee_name=current_user.name,
        )

    data = [FinalResultOut.model_validate(i).model_dump() for i in items]
    return ResponseModel(data=data)


# ---- 更新评定等级 ----
@router.put("/{result_id}/rating", response_model=ResponseModel)
async def set_rating(
    result_id: int,
    body: RatingUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Employee = Depends(require_roles([ROLE_ADMIN])),
):
    """管理员为员工设置评定等级"""
    if body.rating not in RATING_LEVELS:
        raise HTTPException(
            status_code=400,
            detail=f"评定等级必须为：{', '.join(RATING_LEVELS)}"
        )
    try:
        record = await update_rating(db, result_id, body.rating)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    data = FinalResultOut.model_validate(record).model_dump()
    return ResponseModel(message="设置成功", data=data)


# ---- 更新领导评语 ----
@router.put("/{result_id}/comment", response_model=ResponseModel)
async def set_comment(
    result_id: int,
    body: LeaderCommentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    """管理员或领导编辑领导评语"""
    if current_user.role not in (ROLE_ADMIN, ROLE_LEADER):
        raise HTTPException(status_code=403, detail="权限不足")
    try:
        record = await update_leader_comment(db, result_id, body.leader_comment)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    data = FinalResultOut.model_validate(record).model_dump()
    return ResponseModel(message="保存成功", data=data)


# ---- 导出成绩总表 Excel ----
@router.get("/export", response_model=None)
async def export_excel(
    db: AsyncSession = Depends(get_db),
    current_user: Employee = Depends(require_roles([ROLE_ADMIN])),
):
    """导出最终考核成绩总表。按考核类型分Sheet，统一 11 列 + 评语。"""
    cycle = await _get_active_cycle(db)
    all_results = await get_final_results(db, cycle.id)

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # 移除默认Sheet

    headers = [
        "姓名", "部门", "组/中心", "岗位", "岗级", "考核类型",
        "工作积分", "经济指标", "重点任务",
        "加减分", "总分", "评语",
    ]

    # 按考核类型分组
    type_groups = {}
    for r in all_results:
        type_groups.setdefault(r.assess_type, []).append(r)

    for assess_type, items in type_groups.items():
        ws = wb.create_sheet(title=assess_type[:31])  # Sheet名称最长31字符
        ws.append(headers)
        for r in items:
            ws.append([
                r.employee_name, r.department, r.group_name or "",
                r.position or "", r.grade or "", r.assess_type,
                float(r.work_score), float(r.economic_score),
                float(r.key_task_score),
                float(r.bonus_score), float(r.total_score),
                r.leader_comment or "",
            ])

        # 设置列宽
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len * 2 + 2, 12)

    if not wb.sheetnames:
        wb.create_sheet("空")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=final_results_{cycle.name}.xlsx"},
    )


# ---- 导出全部报表（4个Sheet） ----
@router.get("/export-all", response_model=None)
async def export_all_reports(
    db: AsyncSession = Depends(get_db),
    current_user: Employee = Depends(require_roles([ROLE_ADMIN])),
):
    """
    一次性导出四张报表到一个Excel文件的四个Sheet中：
    1. 积分统计表
    2. 综合测评表
    3. 经济指标核算表
    4. 最终考核成绩总表
    """
    cycle = await _get_active_cycle(db)

    from openpyxl import Workbook
    from app.services.score_service import export_score_data
    from app.services.economic_service import calculate_economic_indicators
    from app.models.evaluation import EvalSummary

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Sheet1: 积分统计表 ----
    score_data = await export_score_data(db, cycle.id)
    ws1 = wb.create_sheet("积分统计表")
    ws1.append([
        "员工姓名", "部门", "考核类型",
        "项目积分合计", "公共积分合计", "转型积分合计",
        "总积分", "开方归一化得分",
    ])
    for s in score_data["summaries"]:
        ws1.append([
            s.employee_name, s.department, s.assess_type,
            float(s.project_score_total), float(s.public_score_total),
            float(s.transform_score_total), float(s.total_score),
            float(s.normalized_score),
        ])

    # ---- Sheet2: 综合测评表 ----
    eval_result = await db.execute(
        select(EvalSummary).where(EvalSummary.cycle_id == cycle.id)
    )
    eval_summaries = eval_result.scalars().all()

    ws2 = wb.create_sheet("综合测评表")
    ws2.append([
        "所属部门", "被测评人", "岗位", "考核类型",
        "同事1评分", "同事2评分", "同事3评分", "同事4评分",
        "上级领导评分", "部门领导评分",
        "加权汇总得分", "最终得分(/30)",
    ])
    for e in eval_summaries:
        ws2.append([
            e.department, e.employee_name, e.position or "", e.assess_type,
            float(e.colleague1_score), float(e.colleague2_score),
            float(e.colleague3_score), float(e.colleague4_score),
            float(e.superior_score), float(e.dept_leader_score),
            float(e.weighted_total), float(e.final_score),
        ])

    # ---- Sheet3: 经济指标核算表 ----
    economic_details = await calculate_economic_indicators(db, cycle.id)
    ws3 = wb.create_sheet("经济指标核算表")
    ws3.append([
        "员工姓名", "部门", "组/中心", "岗级", "考核类型",
        "项目名称", "指标类型", "原始值（万元）",
        "参与系数", "完成值（万元）", "目标值（万元）",
        "指标系数", "满分", "得分",
    ])
    for d in economic_details:
        ws3.append([
            d["employee_name"], d["department"], d["group_name"],
            d["grade"], d["assess_type"],
            d["project_name"], d["indicator_type"],
            d["raw_value"], d["participation_coeff"],
            d["completed_value"], d["target_value"],
            d["indicator_coeff"], d["full_mark"], d["score"],
        ])

    # ---- Sheet4: 最终考核成绩总表 ----
    all_results = await get_final_results(db, cycle.id)
    ws4 = wb.create_sheet("最终考核成绩总表")
    ws4.append([
        "姓名", "部门", "组/中心", "岗位", "岗级", "考核类型",
        "工作积分", "经济指标", "重点任务",
        "加减分", "总分", "评语",
    ])
    for r in all_results:
        ws4.append([
            r.employee_name, r.department, r.group_name or "",
            r.position or "", r.grade or "", r.assess_type,
            float(r.work_score), float(r.economic_score),
            float(r.key_task_score),
            float(r.bonus_score), float(r.total_score),
            r.leader_comment or "",
        ])

    # 设置所有Sheet的列宽
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len * 2 + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=all_reports_{cycle.name}.xlsx"},
    )


# ---- 确认考核完成（归档） ----
@router.post("/confirm", response_model=ResponseModel)
async def confirm_complete(
    db: AsyncSession = Depends(get_db),
    current_user: Employee = Depends(require_roles([ROLE_ADMIN])),
):
    """确认本周期考核完成，将周期标记为归档"""
    cycle = await _get_active_cycle(db)
    cycle.is_archived = True
    await db.flush()
    return ResponseModel(message=f"考核周期 {cycle.name} 已确认完成并归档")
