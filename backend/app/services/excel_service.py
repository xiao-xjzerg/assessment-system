"""Excel 导入解析与模板生成"""
import io
from typing import BinaryIO

from openpyxl import Workbook, load_workbook

# ---- 员工 Excel 列映射 ----
EMPLOYEE_COLUMNS = [
    ("姓名", "name"),
    ("部门", "department"),
    ("组/中心", "group_name"),
    ("岗位", "position"),
    ("岗级", "grade"),
    ("联系方式", "phone"),
    ("角色", "role"),
    ("考核类型", "assess_type"),
    ("第二考核类型", "assess_type_secondary"),
]

# ---- 项目 Excel 列映射 ----
PROJECT_COLUMNS = [
    ("项目令号", "project_code"),
    ("项目名称", "project_name"),
    ("项目状态", "project_status"),
    ("项目类型", "project_type"),
    ("实施方式", "impl_method"),
    ("主承部门", "department"),
    ("客户名称", "customer_name"),
    ("合同金额(万元)", "contract_amount"),
    ("项目利润(万元)", "project_profit"),
    ("自研收入(万元)", "self_dev_income"),
    ("产品合同金额(万元)", "product_contract_amount"),
    ("售前活动进度系数", "presale_progress"),
    ("交付活动进度系数", "delivery_progress"),
    ("项目经理", "pm_name"),
    ("签约概率", "signing_probability"),
]


def parse_excel(file: BinaryIO, columns: list[tuple[str, str]]) -> list[dict]:
    """解析 Excel 文件，返回字典列表。columns 为 (表头中文, 字段名) 映射。"""
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    # 读取表头行，建立列索引到字段名的映射
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {col_name: field for col_name, field in columns}

    col_field_map = {}
    for col_idx, cell_value in enumerate(header_row):
        if cell_value and str(cell_value).strip() in header_map:
            col_field_map[col_idx] = header_map[str(cell_value).strip()]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {}
        for col_idx, field in col_field_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[field] = str(val).strip() if val is not None else None
        rows.append(record)

    wb.close()
    return rows


def parse_employee_excel(file: BinaryIO) -> list[dict]:
    return parse_excel(file, EMPLOYEE_COLUMNS)


def parse_project_excel(file: BinaryIO) -> list[dict]:
    return parse_excel(file, PROJECT_COLUMNS)


def generate_template(columns: list[tuple[str, str]], sample_data: list[dict] | None = None) -> bytes:
    """生成 Excel 模板文件，返回 bytes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"

    # 写表头
    headers = [col[0] for col in columns]
    ws.append(headers)

    # 设置列宽
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = max(len(header) * 2 + 4, 15)

    # 写示例数据
    if sample_data:
        field_names = [col[1] for col in columns]
        for row_data in sample_data:
            row = [row_data.get(field, "") for field in field_names]
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_employee_template() -> bytes:
    sample = [
        {
            "name": "张三",
            "department": "实施交付部",
            "group_name": "应用集成中心",
            "position": "项目经理",
            "grade": "T5",
            "phone": "13800138001",
            "role": "项目经理",
            "assess_type": "业务人员",
            "assess_type_secondary": "",
        }
    ]
    return generate_template(EMPLOYEE_COLUMNS, sample)


def generate_project_template() -> bytes:
    sample = [
        {
            "project_code": "PJ2026001",
            "project_name": "示例项目",
            "project_status": "进行中",
            "project_type": "集成",
            "impl_method": "服务",
            "department": "实施交付部",
            "customer_name": "示例客户",
            "contract_amount": "100",
            "project_profit": "30",
            "self_dev_income": "0",
            "product_contract_amount": "0",
            "presale_progress": "0.5",
            "delivery_progress": "0.3",
            "pm_name": "张三",
            "signing_probability": "1",
        }
    ]
    return generate_template(PROJECT_COLUMNS, sample)
